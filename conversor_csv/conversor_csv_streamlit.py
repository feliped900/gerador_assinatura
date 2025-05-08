import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import numbers

st.title("Conversor CSV para Excel com Tratamento de Valor (R$)")

# Upload do arquivo
uploaded_file = st.file_uploader("Faça upload do arquivo CSV", type=["csv"])

if uploaded_file:
    try:
        # Ler CSV com separador ; e encoding latino
        df = pd.read_csv(uploaded_file, sep=';', encoding='latin1', on_bad_lines='skip')

        # Corrigir nomes das colunas
        df.columns = [col.encode('latin1').decode('utf-8').strip() for col in df.columns]

        # Exibir prévia dos dados
        st.subheader("Prévia dos Dados Originais")
        st.dataframe(df.head())

        # Tratar VALOR CHURN
        df['VALOR CHURN'] = df['VALOR CHURN'].astype(str)
        df['VALOR CHURN'] = (
            df['VALOR CHURN']
            .str.replace(r'[^\d,]', '', regex=True)
            .str.replace(',', '.', regex=False)
            .astype(float)
        )

        # Exportar para Excel em memória
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Dados Tratados")
            writer.book.active = writer.sheets["Dados Tratados"]

        # Abrir planilha com openpyxl para aplicar formatação
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active

        # Descobrir a coluna "VALOR CHURN"
        col_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "VALOR CHURN":
                col_idx = idx
                break

        # Aplicar formatação R$ nas células
        if col_idx:
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = 'R$ #,##0.00'

        # Salvar para download
        output_final = BytesIO()
        wb.save(output_final)
        output_final.seek(0)

        st.success("Arquivo tratado com sucesso!")

        # Botão de download
        st.download_button(
            label="📥 Baixar Excel Tratado",
            data=output_final,
            file_name="cancelamentos_tratado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Ocorreu um erro ao processar o arquivo: {e}")
